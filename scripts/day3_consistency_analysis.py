#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError:  # optional
    Workbook = None
    load_workbook = None

SENTIMENT_LABELS = ["POS", "NEG", "NEU"]
BIO_TAG_PATTERN = re.compile(r"\(([BIO](?:-[A-Za-z]+)?)\)")
ENTITY_PATTERN = re.compile(r"(PER|ORG|LOC|TIME)\s*[:：]\s*([^,，;；|]+)")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("input_file", type=Path)
    p.add_argument("--output-dir", type=Path, default=Path("reports"))
    p.add_argument("--id-col", default="ID")
    p.add_argument("--task-col", default="任务类型")
    p.add_argument("--text-col", default="原始文本")
    p.add_argument("--label-col", default="标注结果")
    return p.parse_args()


def load_records(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    if path.suffix.lower() in {".xlsx", ".xls"}:
        if load_workbook is None:
            raise RuntimeError("读取 Excel 需要安装 openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h is not None else "" for h in rows[0]]
        out = []
        for r in rows[1:]:
            out.append(
                {
                    headers[i]: ("" if r[i] is None else str(r[i]))
                    for i in range(len(headers))
                }
            )
        return out
    raise ValueError("仅支持 CSV / Excel")


def extract_sentiment_label(text: str) -> str | None:
    upper = text.upper()
    for lb in SENTIMENT_LABELS:
        if re.search(rf"\b{lb}\b", upper):
            return lb
    return None


def extract_entities(text: str) -> tuple[tuple[str, str], ...]:
    entities = ENTITY_PATTERN.findall(text)
    return tuple(sorted((k, v.strip()) for k, v in entities))


def extract_bio_tags(text: str) -> list[str]:
    return BIO_TAG_PATTERN.findall(text)


def validate_bio(tags: list[str]) -> tuple[bool, str]:
    prev_prefix, prev_type = "O", None
    for i, tag in enumerate(tags):
        if tag == "O":
            prev_prefix, prev_type = "O", None
            continue
        prefix, typ = (tag.split("-", 1) + [None])[:2] if "-" in tag else (tag, None)
        if prefix == "I":
            if i == 0:
                return False, "I 标签不能开头"
            if prev_prefix not in {"B", "I"}:
                return False, "I 前置标签非法"
            if prev_type != typ:
                return False, "I 标签类型与前一标签不一致"
        elif prefix != "B":
            return False, f"非法前缀: {prefix}"
        prev_prefix, prev_type = prefix, typ
    return True, "合法"


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def main() -> None:
    a = parse_args()
    a.output_dir.mkdir(parents=True, exist_ok=True)
    records = load_records(a.input_file)
    for col in [a.id_col, a.task_col, a.text_col, a.label_col]:
        if any(col not in r for r in records):
            raise ValueError(f"缺少必要列: {col}")

    total = len(records)
    missing_by_col = Counter()
    id_counter, text_counter = Counter(), Counter()
    sentiments = Counter()
    entities_by_text = defaultdict(list)
    bio_issues = []
    anomaly_list = []

    for row in records:
        for col, val in row.items():
            if val is None or str(val).strip() == "":
                missing_by_col[col] += 1
        id_counter[str(row[a.id_col])] += 1
        text_counter[str(row[a.text_col])] += 1

        label = str(row.get(a.label_col, ""))
        sentiment = extract_sentiment_label(label)
        if sentiment:
            sentiments[sentiment] += 1

        ents = extract_entities(label)
        if ents:
            entities_by_text[str(row[a.text_col])].append((str(row[a.id_col]), ents))

        bio_tags = extract_bio_tags(label)
        if bio_tags:
            ok, reason = validate_bio(bio_tags)
            if not ok:
                bio_issues.append(
                    {
                        "id": row[a.id_col],
                        "text": row[a.text_col],
                        "issue": "BIO 序列不合法",
                        "reason": reason,
                    }
                )

    dup_id = {k for k, v in id_counter.items() if v > 1}
    dup_text = {k for k, v in text_counter.items() if v > 1}

    for row in records:
        reasons = []
        if str(row[a.id_col]) in dup_id:
            reasons.append("重复ID")
        if str(row[a.text_col]) in dup_text:
            reasons.append("重复文本")
        if any((v is None or str(v).strip() == "") for v in row.values()):
            reasons.append("存在缺失值")
        if reasons:
            anomaly_list.append(
                {
                    "id": row[a.id_col],
                    "task_type": row[a.task_col],
                    "text": row[a.text_col],
                    "issue": "；".join(reasons),
                }
            )

    ner_inconsistency = []
    for text, rows in entities_by_text.items():
        uniq = {tuple(sorted(ent)) for _, ent in rows}
        if len(uniq) > 1:
            ner_inconsistency.append(
                {
                    "text": text,
                    "sample_ids": ", ".join(i for i, _ in rows),
                    "entity_variants": " | ".join(str(u) for u in sorted(uniq)),
                    "issue": "同文本存在不同实体标注",
                }
            )

    summary = [
        {"metric": "total_samples", "value": total},
        {"metric": "missing_cells", "value": sum(missing_by_col.values())},
        {
            "metric": "duplicate_id_samples",
            "value": sum(1 for v in id_counter.values() if v > 1),
        },
        {
            "metric": "duplicate_text_samples",
            "value": sum(1 for v in text_counter.values() if v > 1),
        },
        {"metric": "ner_inconsistency_cases", "value": len(ner_inconsistency)},
        {"metric": "bio_invalid_cases", "value": len(bio_issues)},
    ]
    missing_stats = [
        {
            "column": c,
            "missing_count": n,
            "missing_ratio": round(n / total, 4) if total else 0,
        }
        for c, n in missing_by_col.items()
    ]
    duplicate_stats = [
        {
            "metric": "duplicate_id",
            "count": sum(1 for v in id_counter.values() if v > 1),
            "ratio": round((sum(1 for v in id_counter.values() if v > 1) / total), 4)
            if total
            else 0,
        },
        {
            "metric": "duplicate_text",
            "count": sum(1 for v in text_counter.values() if v > 1),
            "ratio": round((sum(1 for v in text_counter.values() if v > 1) / total), 4)
            if total
            else 0,
        },
    ]
    sentiment_distribution = [
        {
            "label": lb,
            "count": cnt,
            "ratio": round(cnt / sum(sentiments.values()), 4) if sentiments else 0,
        }
        for lb, cnt in sentiments.items()
    ]

    csv_dir = a.output_dir / "day3_consistency_analysis"
    csv_dir.mkdir(parents=True, exist_ok=True)
    write_csv(csv_dir / "summary.csv", summary)
    write_csv(csv_dir / "missing_stats.csv", missing_stats)
    write_csv(csv_dir / "duplicate_stats.csv", duplicate_stats)
    write_csv(csv_dir / "sentiment_distribution.csv", sentiment_distribution)
    write_csv(csv_dir / "ner_inconsistency.csv", ner_inconsistency)
    write_csv(csv_dir / "bio_issues.csv", bio_issues)
    write_csv(csv_dir / "anomaly_list.csv", anomaly_list)

    excel_path = csv_dir / "consistency_analysis_report.xlsx"
    if Workbook is not None:
        wb = Workbook()
        wb.active.title = "summary"
        sheets = {
            "summary": summary,
            "missing_stats": missing_stats,
            "duplicate_stats": duplicate_stats,
            "sentiment_distribution": sentiment_distribution,
            "ner_inconsistency": ner_inconsistency,
            "bio_issues": bio_issues,
            "anomaly_list": anomaly_list,
        }
        for idx, (name, rows) in enumerate(sheets.items()):
            ws = wb.active if idx == 0 else wb.create_sheet(name)
            ws.title = name
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                for r in rows:
                    ws.append([r.get(h, "") for h in headers])
        wb.save(excel_path)

    print(f"分析完成，总样本数: {total}")
    print(f"CSV 输出目录: {csv_dir}")
    if Workbook is not None:
        print(f"Excel 报告: {excel_path}")
    else:
        print("Excel 报告未生成（未安装 openpyxl）")


if __name__ == "__main__":
    # uv run python -m scripts.day3_consistency_analysis data/day02_annotation.csv
    # uv run python -m scripts.day3_consistency_analysis data/day02_annotation.xlsx
    main()
